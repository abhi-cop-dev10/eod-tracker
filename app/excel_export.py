from pathlib import Path
from datetime import date

from PyQt6.QtWidgets import QFileDialog, QMessageBox

from app.database import get_tasks_for_export, get_setting, set_setting

COLUMN_MAP = {
    "task_name":   "A",
    "eod_status":  "B",
    "client_name": "C",
    "start_time":  "D",
    "end_time":    "E",
    "duration":    "F",
    "assigned_to": "G",
    "assigned_by": "H",
    "what_done":   "I",
}

from app.paths import templates_dir as _templates_dir
DEFAULT_TEMPLATE = _templates_dir() / "eod_template.xlsx"


def _col_letter_to_index(letter):
    """Convert column letter to openpyxl 1-based index."""
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def export_excel(parent=None):
    try:
        import openpyxl
    except ImportError:
        _error(parent, "openpyxl is not installed.\nRun: pip install openpyxl")
        return

    # Determine template path
    template_path_str = get_setting('excel_template_path', '')
    template_path = Path(template_path_str) if template_path_str else DEFAULT_TEMPLATE

    if not template_path.exists():
        # If no template, offer to create a basic one
        reply = QMessageBox.question(
            parent,
            "Template Not Found",
            f"Template not found at:\n{template_path}\n\nCreate a basic template and export?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            _create_basic_template(template_path)
        else:
            return

    # Load workbook
    try:
        wb = openpyxl.load_workbook(str(template_path), keep_vba=False)
    except Exception as e:
        _error(parent, f"Failed to open template:\n{e}")
        return

    ws = wb.active

    # Read settings
    data_start_row = int(get_setting('data_start_row', '2'))

    # Load column map from settings if customised (validate A-Z / AA-ZZ / AAA-ZZZ)
    import re
    _COL_RE = re.compile(r'^[A-Z]{1,3}$')
    col_map = dict(COLUMN_MAP)
    for field in COLUMN_MAP:
        saved = get_setting(f'col_map_{field}', '').strip().upper()
        if saved and _COL_RE.match(saved):
            col_map[field] = saved

    # Get today's tasks
    tasks = get_tasks_for_export()
    if not tasks:
        QMessageBox.information(
            parent,
            "No Tasks",
            "No tasks recorded for today. Add and start some tasks first.",
        )
        return

    # Write data
    for i, task in enumerate(tasks):
        row_num = data_start_row + i
        for field, col_letter in col_map.items():
            col_idx = _col_letter_to_index(col_letter)
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = task.get(field, '') or ''

    # Determine employee name for filename
    employee_name = get_setting('employee_name', 'Employee')
    safe_name = "".join(c for c in employee_name if c.isalnum() or c in (' ', '_', '-')).strip()
    today_str = date.today().strftime("%Y-%m-%d")
    default_filename = f"EOD_{safe_name}_{today_str}.xlsx"

    # Save dialog
    save_path, _ = QFileDialog.getSaveFileName(
        parent,
        "Save EOD Report",
        str(Path.home() / "Desktop" / default_filename),
        "Excel Files (*.xlsx)"
    )
    if not save_path:
        return

    try:
        wb.save(save_path)
    except Exception as e:
        _error(parent, f"Failed to save file:\n{e}")
        return

    from app.toast import show_toast
    show_toast("Excel saved!")

    # Offer to open the file
    reply = QMessageBox.question(
        parent, "Saved",
        f"Report saved to:\n{save_path}\n\nOpen the file now?",
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
    )
    if reply == QMessageBox.StandardButton.Yes:
        import os
        os.startfile(save_path)


def _error(parent, msg):
    QMessageBox.critical(parent, "Export Error", msg)


def _create_basic_template(path: Path):
    """Create a minimal xlsx template with header row."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EOD Report"

        headers = [
            "Task Name", "Status", "Client", "Start Time",
            "End Time", "Duration", "Assigned To", "Assigned By", "What Have Done?"
        ]

        header_font  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill(fill_type="solid", fgColor="E94560")
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border

        col_widths = [30, 15, 20, 12, 12, 12, 20, 20, 40]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        ws.row_dimensions[1].height = 22
        wb.save(str(path))
    except Exception:
        pass
